from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ===== 可調整設定 =====
INPUT_FILE = Path("ICP_Audit_Report_20260707_111225.xlsx")
OUTPUT_FILE = Path("ICP_Audit_高風險案件_文字閱讀版_配色一致.xlsx")
SOURCE_SHEET = "Sheet1"
FILTER_COLUMN = "LLM研判等級"
TARGET_LEVELS = {"High", "Medium"}

# 與原本輸出成品一致的色碼
COLORS = {
    "navy": "16324F",          # 主標題、清單表頭
    "teal": "245B78",          # 案件標題
    "pale_blue": "EAF2F8",     # 副標題、LLM 推理區
    "pale_teal": "E8F5F5",     # 實體區塊標題
    "pale_yellow": "FFF7E3",   # 摘要指標、黑名單資訊區
    "pale_gray": "F3F5F7",     # 一般欄位標籤
    "white": "FFFFFF",
    "text": "23313F",
    "muted": "607080",
}


def s(value: Any, empty: str = "—") -> str:
    if value is None or str(value).strip() == "":
        return empty
    return str(value).strip()


def solid(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def merge_write(ws, cell_range: str, value: Any) -> None:
    ws.merge_cells(cell_range)
    ws[cell_range.split(":")[0]] = value


def style_cell(cell, *, fill=None, color=None, size=10, bold=False,
               horizontal="left", vertical="center", wrap=True):
    if fill:
        cell.fill = solid(fill)
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color or COLORS["text"])
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def style_range(ws, cell_range: str, **kwargs) -> None:
    for row in ws[cell_range]:
        for cell in row:
            style_cell(cell, **kwargs)


def read_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("來源工作表沒有資料")
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    if FILTER_COLUMN not in headers:
        raise ValueError(f"找不到欄位：{FILTER_COLUMN}")
    idx = headers.index(FILTER_COLUMN)
    records = []
    for row in rows[1:]:
        if idx < len(row) and str(row[idx]).strip() in TARGET_LEVELS:
            records.append({h: (row[i] if i < len(row) else None) for i, h in enumerate(headers)})
    return headers, records


def set_card_columns(ws):
    widths = {"A": 16, "B": 26, "C": 26, "D": 3, "E": 16, "F": 26, "G": 26, "H": 3}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_reading_sheet(ws, records):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    set_card_columns(ws)

    merge_write(ws, "A1:H1", "LLM 高與中風險案件｜審計閱讀版")
    style_range(ws, "A1:H1", fill=COLORS["navy"], color=COLORS["white"], size=18,
                bold=True, horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 42

    merge_write(ws, "A2:H2", f"共 {len(records)} 筆 High/Medium 風險案件｜完整保留所有原始欄位，改以案件卡片方式呈現")
    style_range(ws, "A2:H2", fill=COLORS["pale_blue"], color=COLORS["muted"], size=10,
                horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 26

    row = 4
    for n, rec in enumerate(records, start=1):
        query_name = s(rec.get("查詢名稱"))
        watch_name = s(rec.get("黑名單名稱"))
        merge_write(ws, f"A{row}:H{row}", f"案件 {n:02d}　{query_name}  ⇄  {watch_name}")
        style_range(ws, f"A{row}:H{row}", fill=COLORS["teal"], color=COLORS["white"],
                    size=13, bold=True, vertical="center")
        ws.row_dimensions[row].height = 32
        row += 1

        # 摘要列
        summary = [
            ("A", "條件 ID", "B", s(rec.get("條件ID"))),
            ("C", "原 XML 命中率", "D", s(rec.get("原XML命中率"))),
            ("E", "LLM 研判等級", "F", s(rec.get("LLM研判等級"))),
            ("G", "風險判定依據", "H", s(rec.get("風險判定依據"))),
        ]
        for lc, label, vc, value in summary:
            ws[f"{lc}{row}"] = label
            ws[f"{vc}{row}"] = value
            style_cell(ws[f"{lc}{row}"], fill=COLORS["pale_yellow"], size=9, bold=True)
            style_cell(ws[f"{vc}{row}"], fill=COLORS["white"], color=COLORS["muted"], size=10,
                       horizontal="center")
        ws.row_dimensions[row].height = 26
        row += 1

        ws[f"A{row}"] = "來源檔案"
        merge_write(ws, f"B{row}:H{row}", s(rec.get("來源檔案")))
        style_cell(ws[f"A{row}"], fill=COLORS["pale_gray"], size=11, bold=True)
        style_range(ws, f"B{row}:H{row}", fill=COLORS["white"], color=COLORS["muted"], size=9)
        ws.row_dimensions[row].height = 28
        row += 1

        merge_write(ws, f"A{row}:C{row}", "查詢實體")
        merge_write(ws, f"E{row}:G{row}", "黑名單實體")
        style_range(ws, f"A{row}:C{row}", fill=COLORS["pale_teal"], color=COLORS["teal"],
                    size=11, bold=True)
        style_range(ws, f"E{row}:G{row}", fill=COLORS["pale_teal"], color=COLORS["teal"],
                    size=11, bold=True)
        ws.row_dimensions[row].height = 26
        row += 1

        # 名稱
        ws[f"A{row}"] = "名稱"
        merge_write(ws, f"B{row}:C{row}", query_name)
        ws[f"E{row}"] = "名稱"
        merge_write(ws, f"F{row}:G{row}", watch_name)
        for c in (f"A{row}", f"E{row}"):
            style_cell(ws[c], fill=COLORS["pale_gray"], size=9, bold=True)
        style_range(ws, f"B{row}:C{row}", fill=COLORS["white"], size=10)
        style_range(ws, f"F{row}:G{row}", fill=COLORS["white"], size=10)
        ws.row_dimensions[row].height = 32
        row += 1

        # 國家 / 黑名單 ID
        ws[f"A{row}"] = "國家"
        merge_write(ws, f"B{row}:C{row}", s(rec.get("查詢國家")))
        ws[f"E{row}"] = "黑名單 ID"
        merge_write(ws, f"F{row}:G{row}", s(rec.get("黑名單ID")))
        for c in (f"A{row}", f"E{row}"):
            style_cell(ws[c], fill=COLORS["pale_gray"], size=9, bold=True)
        style_range(ws, f"B{row}:C{row}", fill=COLORS["white"], size=10)
        style_range(ws, f"F{row}:G{row}", fill=COLORS["white"], size=10)
        ws.row_dimensions[row].height = 32
        row += 1

        # 城市 / 黑名單地址
        ws[f"A{row}"] = "城市"
        merge_write(ws, f"B{row}:C{row}", s(rec.get("查詢城市")))
        ws[f"E{row}"] = "地址"
        merge_write(ws, f"F{row}:G{row}", s(rec.get("黑名單地址")))
        for c in (f"A{row}", f"E{row}"):
            style_cell(ws[c], fill=COLORS["pale_gray"], size=9, bold=True, vertical="top")
        style_range(ws, f"B{row}:C{row}", fill=COLORS["white"], size=10, vertical="top")
        style_range(ws, f"F{row}:G{row}", fill=COLORS["white"], size=10, vertical="top")
        ws.row_dimensions[row].height = 88
        row += 1

        # 查詢地址
        ws[f"A{row}"] = "地址"
        merge_write(ws, f"B{row}:C{row}", s(rec.get("查詢地址")))
        style_cell(ws[f"A{row}"], fill=COLORS["pale_gray"], size=9, bold=True, vertical="top")
        style_range(ws, f"B{row}:C{row}", fill=COLORS["white"], size=10, vertical="top")
        ws.row_dimensions[row].height = 52
        row += 1

        # 黑名單完整資訊
        ws[f"A{row}"] = "黑名單完整資訊"
        merge_write(ws, f"B{row}:H{row}", s(rec.get("黑名單完整資訊")))
        style_cell(ws[f"A{row}"], fill=COLORS["pale_yellow"], size=9, bold=True, vertical="top")
        style_range(ws, f"B{row}:H{row}", fill=COLORS["white"], color=COLORS["muted"], size=9,
                    vertical="top")
        ws.row_dimensions[row].height = 115
        row += 1

        # LLM 推理
        ws[f"A{row}"] = "LLM 分析推理理由"
        merge_write(ws, f"B{row}:H{row}", s(rec.get("LLM分析推理理由")))
        style_cell(ws[f"A{row}"], fill=COLORS["pale_blue"], color=COLORS["teal"], size=9,
                   bold=True, vertical="top")
        style_range(ws, f"B{row}:H{row}", fill=COLORS["white"], color=COLORS["text"], size=9,
                    vertical="top")
        ws.row_dimensions[row].height = 105
        row += 1

        ws.row_dimensions[row].height = 14
        row += 1

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:2"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


def build_detail_sheet(ws, headers, records):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    for r, rec in enumerate(records, 2):
        for c, header in enumerate(headers, 1):
            ws.cell(r, c, rec.get(header))

    last_col = get_column_letter(len(headers))
    last_row = len(records) + 1

    style_range(ws, f"A1:{last_col}1", fill=COLORS["navy"], color=COLORS["white"], size=11,
                bold=True, horizontal="center")
    ws.row_dimensions[1].height = 30
    style_range(ws, f"A2:{last_col}{last_row}", color=COLORS["text"], size=10, vertical="top")

    wide = {"來源檔案", "查詢名稱", "查詢地址", "黑名單名稱", "黑名單地址", "黑名單完整資訊", "LLM分析推理理由"}
    narrow = {"條件ID", "查詢國家", "查詢城市", "黑名單ID", "原XML命中率", "LLM研判等級", "風險判定依據", "LLM判定是否同實體"}
    for i, header in enumerate(headers, 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 38 if header in wide else 16 if header in narrow else 20
    for r in range(2, last_row + 1):
        ws.row_dimensions[r].height = 72

    table = Table(displayName="SameEntityTable", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                          showLastColumn=False, showRowStripes=True,
                                          showColumnStripes=False)
    ws.add_table(table)


def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"找不到輸入檔案：{INPUT_FILE.resolve()}")

    wb = load_workbook(INPUT_FILE)
    if SOURCE_SHEET not in wb.sheetnames:
        raise ValueError(f"找不到工作表：{SOURCE_SHEET}")

    headers, records = read_records(wb[SOURCE_SHEET])
    if not records:
        raise ValueError(f"沒有找到 {FILTER_COLUMN} = {FILTER_VALUE} 的資料")

    for name in ["同實體完整清單", "同實體閱讀版"]:
        if name in wb.sheetnames:
            del wb[name]

    detail_ws = wb.create_sheet("同實體完整清單")
    reading_ws = wb.create_sheet("同實體閱讀版")
    build_detail_sheet(detail_ws, headers, records)
    build_reading_sheet(reading_ws, records)

    wb.save(OUTPUT_FILE)
    print(f"完成：{OUTPUT_FILE.resolve()}")
    print(f"篩選筆數：{len(records)}")


if __name__ == "__main__":
    main()
