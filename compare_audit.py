import os
import glob
import re
import json
from pathlib import Path
from typing import Any
from bs4 import BeautifulSoup
import pandas as pd
from openai import OpenAI
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# 1. 初始化 OpenAI 客戶端，連接您的本地 vLLM (gemma-4:31B)
client = OpenAI(
    base_url="http://192.168.39.143:8002/v1",
    api_key="empty_api_key_for_vllm"
)

# 本地模型 ID 與併發限制線程數
MODEL_NAME = "gemma-4:31B"
MAX_WORKERS = 16

# Excel 報表配色系
COLORS = {
    "navy": "16324F",          # 主標題、清單表頭
    "teal": "245B78",          # 案件標題
    "pale_blue": "EAF2F8",     # 副標題、LLM 推理區
    "pale_teal": "E8F5F5",     # 實體區塊標題
    "pale_yellow": "FFF7E3",   # 摘要指標、黑名單資訊區
    "pale_gray": "F3F5F7",     # 一般欄位標籤
    "white": "FFFFFF",
    "text": "23313F",
    "muted": "607080",
}

def s(value: Any, empty: str = "—") -> str:
    if value is None or str(value).strip() == "":
        return empty
    return str(value).strip()

def solid(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)

def merge_write(ws, cell_range: str, value: Any) -> None:
    ws.merge_cells(cell_range)
    ws[cell_range.split(":")[0]] = value

def style_cell(cell, *, fill=None, color=None, size=10, bold=False,
               horizontal="left", vertical="center", wrap=True):
    if fill:
        cell.fill = solid(fill)
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color or COLORS["text"])
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def style_range(ws, cell_range: str, **kwargs) -> None:
    for row in ws[cell_range]:
        for cell in row:
            style_cell(cell, **kwargs)

def set_card_columns(ws):
    widths = {"A": 16, "B": 26, "C": 26, "D": 10, "E": 16, "F": 26, "G": 26, "H": 24}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def build_reading_sheet(ws, records):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    set_card_columns(ws)

    merge_write(ws, "A1:H1", "LLM 高與中風險案件｜審計閱讀版")
    style_range(ws, "A1:H1", fill=COLORS["navy"], color=COLORS["white"], size=18,
                bold=True, horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 42

    merge_write(ws, "A2:H2", f"共 {len(records)} 筆 High/Medium 風險案件｜完整保留所有欄位，以案件卡片方式呈現")
    style_range(ws, "A2:H2", fill=COLORS["pale_blue"], color=COLORS["muted"], size=10,
                horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 26

    row = 4
    for n, rec in enumerate(records, start=1):
        query_name = s(rec.get("查詢名稱"))
        watch_name = s(rec.get("黑名單名稱"))
        merge_write(ws, f"A{row}:H{row}", f"案件 {n:02d}　{query_name}  ⇄  {watch_name}")
        style_range(ws, f"A{row}:H{row}", fill=COLORS["teal"], color=COLORS["white"],
                    size=13, bold=True, vertical="center")
        ws.row_dimensions[row].height = 32
        row += 1

        # 摘要列
        summary = [
            ("A", "條件 ID", "B", s(rec.get("條件ID"))),
            ("C", "原 XML 命中率", "D", s(rec.get("原XML命中率"))),
            ("E", "LLM 研判等級", "F", s(rec.get("LLM研判等級"))),
            ("G", "風險判定依據", "H", s(rec.get("風險判定依據"))),
        ]
        for lc, label, vc, value in summary:
            ws[f"{lc}{row}"] = label
            ws[f"{vc}{row}"] = value
            style_cell(ws[f"{lc}{row}"], fill=COLORS["pale_yellow"], size=9, bold=True)
            style_cell(ws[f"{vc}{row}"], fill=COLORS["white"], color=COLORS["muted"], size=10,
                       horizontal="center")
        ws.row_dimensions[row].height = 26
        row += 1

        ws[f"A{row}"] = "來源檔案"
        merge_write(ws, f"B{row}:H{row}", s(rec.get("來源檔案")))
        style_cell(ws[f"A{row}"], fill=COLORS["pale_gray"], size=11, bold=True)
        style_range(ws, f"B{row}:H{row}", fill=COLORS["white"], color=COLORS["muted"], size=9)
        ws.row_dimensions[row].height = 28
        row += 1

        merge_write(ws, f"A{row}:C{row}", "查詢實體")
        merge_write(ws, f"E{row}:G{row}", "黑名單實體")
        style_range(ws, f"A{row}:C{row}", fill=COLORS["pale_teal"], color=COLORS["teal"],
                    size=11, bold=True)
        style_range(ws, f"E{row}:G{row}", fill=COLORS["pale_teal"], color=COLORS["teal"],
                    size=11, bold=True)
        ws.row_dimensions[row].height = 26
        row += 1

        # 名稱
        ws[f"A{row}"] = "名稱"
        merge_write(ws, f"B{row}:C{row}", query_name)
        ws[f"E{row}"] = "名稱"
        merge_write(ws, f"F{row}:G{row}", watch_name)
        for c in (f"A{row}", f"E{row}"):
            style_cell(ws[c], fill=COLORS["pale_gray"], size=9, bold=True)
        style_range(ws, f"B{row}:C{row}", fill=COLORS["white"], size=10)
        style_range(ws, f"F{row}:G{row}", fill=COLORS["white"], size=10)
        ws.row_dimensions[row].height = 32
        row += 1

        # 國家 / 黑名單 ID
        ws[f"A{row}"] = "國家"
        merge_write(ws, f"B{row}:C{row}", s(rec.get("查詢國家")))
        ws[f"E{row}"] = "黑名單 ID"
        merge_write(ws, f"F{row}:G{row}", s(rec.get("黑名單ID")))
        for c in (f"A{row}", f"E{row}"):
            style_cell(ws[c], fill=COLORS["pale_gray"], size=9, bold=True)
        style_range(ws, f"B{row}:C{row}", fill=COLORS["white"], size=10)
        style_range(ws, f"F{row}:G{row}", fill=COLORS["white"], size=10)
        ws.row_dimensions[row].height = 32
        row += 1

        # 城市 / 黑名單地址
        ws[f"A{row}"] = "城市"
        merge_write(ws, f"B{row}:C{row}", s(rec.get("查詢城市")))
        ws[f"E{row}"] = "地址"
        merge_write(ws, f"F{row}:G{row}", s(rec.get("黑名單地址")))
        for c in (f"A{row}", f"E{row}"):
            style_cell(ws[c], fill=COLORS["pale_gray"], size=9, bold=True, vertical="top")
        style_range(ws, f"B{row}:C{row}", fill=COLORS["white"], size=10, vertical="top")
        style_range(ws, f"F{row}:G{row}", fill=COLORS["white"], size=10, vertical="top")
        ws.row_dimensions[row].height = 88
        row += 1

        # 查詢地址
        ws[f"A{row}"] = "地址"
        merge_write(ws, f"B{row}:C{row}", s(rec.get("查詢地址")))
        style_cell(ws[f"A{row}"], fill=COLORS["pale_gray"], size=9, bold=True, vertical="top")
        style_range(ws, f"B{row}:C{row}", fill=COLORS["white"], size=10, vertical="top")
        ws.row_dimensions[row].height = 52
        row += 1

        # 黑名單完整資訊
        ws[f"A{row}"] = "黑名單完整資訊"
        merge_write(ws, f"B{row}:H{row}", s(rec.get("黑名單完整資訊")))
        style_cell(ws[f"A{row}"], fill=COLORS["pale_yellow"], size=9, bold=True, vertical="top")
        style_range(ws, f"B{row}:H{row}", fill=COLORS["white"], color=COLORS["muted"], size=9,
                    vertical="top")
        ws.row_dimensions[row].height = 115
        row += 1

        # LLM 推理
        ws[f"A{row}"] = "LLM 分析推理理由"
        merge_write(ws, f"B{row}:H{row}", s(rec.get("LLM分析推理理由")))
        style_cell(ws[f"A{row}"], fill=COLORS["pale_blue"], color=COLORS["teal"], size=9,
                   bold=True, vertical="top")
        style_range(ws, f"B{row}:H{row}", fill=COLORS["white"], color=COLORS["text"], size=9,
                    vertical="top")
        ws.row_dimensions[row].height = 105
        row += 1

        ws.row_dimensions[row].height = 14
        row += 1

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:2"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

def build_detail_sheet(ws, headers, records):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    for r, rec in enumerate(records, 2):
        for c, header in enumerate(headers, 1):
            ws.cell(r, c, rec.get(header))

    last_col = get_column_letter(len(headers))
    last_row = len(records) + 1

    style_range(ws, f"A1:{last_col}1", fill=COLORS["navy"], color=COLORS["white"], size=11,
                bold=True, horizontal="center")
    ws.row_dimensions[1].height = 30
    style_range(ws, f"A2:{last_col}{last_row}", color=COLORS["text"], size=10, vertical="top")

    wide = {"來源檔案", "查詢名稱", "查詢地址", "黑名單名稱", "黑名單地址", "黑名單完整資訊", "LLM分析推理理由"}
    narrow = {"條件ID", "查詢國家", "查詢城市", "黑名單ID", "原XML命中率", "LLM研判等級", "風險判定依據"}
    for i, header in enumerate(headers, 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 38 if header in wide else 16 if header in narrow else 20
    for r in range(2, last_row + 1):
        ws.row_dimensions[r].height = 72

    table = Table(displayName="RiskEntityTable", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                          showLastColumn=False, showRowStripes=True,
                                          showColumnStripes=False)
    ws.add_table(table)

def generate_formatted_excel_report(output_excel):
    """
    使用 openpyxl 讀出全審計結果，並新增美化的「高與中風險清單」及卡片式的「高與中風險審計閱讀版」分頁。
    """
    wb = load_workbook(output_excel)
    source_sheet = wb.sheetnames[0]
    ws = wb[source_sheet]
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
        
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    if "LLM研判等級" not in headers:
        return
        
    idx = headers.index("LLM研判等級")
    target_levels = {"High", "Medium"}
    records = []
    for row in rows[1:]:
        if idx < len(row) and str(row[idx]).strip() in target_levels:
            records.append({h: (row[i] if i < len(row) else None) for i, h in enumerate(headers)})
            
    if not records:
        print("[INFO] 無 High / Medium 高中風險案件，跳過建立閱讀版分頁。")
        return
        
    for name in ["高與中風險清單", "高與中風險審計閱讀版"]:
        if name in wb.sheetnames:
            del wb[name]

    detail_ws = wb.create_sheet("高與中風險清單")
    reading_ws = wb.create_sheet("高與中風險審計閱讀版")
    
    build_detail_sheet(detail_ws, headers, records)
    build_reading_sheet(reading_ws, records)
    
    wb.save(output_excel)

def clean_text(text):
    if text is None:
        return ""
    return str(text).strip()

def parse_xml_file(xml_path):
    """
    解析 XML 檔案，過濾出 percentage >= 75.0% 的所有 Result 對照紀錄。
    """
    print(f"正在讀取並解析 XML 檔案: {os.path.basename(xml_path)}...")
    with open(xml_path, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    
    soup = BeautifulSoup(content, 'xml')
    results = soup.find_all('Result')
    
    matched_pairs = []
    
    for r in results:
        condition = r.find('Condition')
        if not condition:
            continue
            
        condition_id = clean_text(condition.find('ConditionId').text if condition.find('ConditionId') else "")
        query_name = clean_text(condition.find('QueryName').text if condition.find('QueryName') else "")
        country = clean_text(condition.find('Country').text if condition.find('Country') else "")
        city = clean_text(condition.find('City').text if condition.find('City') else "")
        address = clean_text(condition.find('Address').text if condition.find('Address') else "")
        
        # 尋找底下的所有 Party 節點
        parties = r.find_all('Party')
        for p in parties:
            # 兼容大小寫 groupid / Groupid 等
            group_id_node = p.find(re.compile('^groupid$', re.I))
            percentage_node = p.find(re.compile('^percentage$', re.I))
            name_node = p.find(re.compile('^name$', re.I))
            addr_node = p.find(re.compile('^address$', re.I))
            content_node = p.find(re.compile('^content$', re.I))
            
            group_id = clean_text(group_id_node.text if group_id_node else "")
            pct_str = clean_text(percentage_node.text if percentage_node else "0")
            party_name = clean_text(name_node.text if name_node else "")
            party_address = clean_text(addr_node.text if addr_node else "")
            party_content = clean_text(content_node.text if content_node else "")
            
            try:
                percentage = float(pct_str)
            except ValueError:
                percentage = 0.0
                
            # 只篩選出命中率 >= 75% 的
            if percentage >= 75.0:
                matched_pairs.append({
                    "condition_id": condition_id,
                    "query_name": query_name,
                    "query_country": country,
                    "query_city": city,
                    "query_address": address,
                    "party_id": group_id,
                    "party_name": party_name,
                    "party_address": party_address,
                    "party_content": party_content,
                    "xml_percentage": percentage
                })
                
    return matched_pairs

def get_llm_judgment(pair):
    """
    呼叫本地 gemma-4:31B 進行深度語意比對。
    """
    prompt = f"""你是一名專業的進出口貿易合規審計專家（ICP/RPS Compliance Auditor）。
我們需要比對一筆「客戶/供應商查詢條件」與一筆「戰略性高科技貨品限制實體（黑名單）資料」，評估兩者是否為同一個實體、關係企業（如分公司、子公司、母公司、轉投資公司等），或是純粹的「字面重疊誤判（False Positive）」。

【查詢條件 (Condition)】
- 條件 ID: {pair['condition_id']}
- 查詢名稱 (QueryName): {pair['query_name']}
- 國家 (Country): {pair['query_country']}
- 城市 (City): {pair['query_city']}
- 查詢地址 (Address): {pair['query_address']}

【黑名單實體 (Party)】
- 限制實體 ID: {pair['party_id']}
- 限制實體名稱 (Name): {pair['party_name']}
- 限制實體地址 (Address): {pair['party_address']}
- 限制實體關聯詳細內容 (Content): {pair['party_content']}

請遵循以下思考框架進行細緻的語意分析：
1. **名稱主體分析 (Name Core Element)**：
   - 移除常見法律後綴詞（如 GMBH, CO. KG, INC, LTD, PTE LTD, LLC 等）。
   - 比對兩者核心名稱（例如 EBM-PAPST 是否與黑名單中的名稱實質關聯）。
   - 注意中英文對譯（例如 廈門算能科技 與 Xiamen Sophgo）。
2. **地址物理一致性與地理常識分析 (Address & Geography)**：
   - 是否在同一棟商業大樓或物流園區？如果是像 Software Park (軟體園)、Midview City、工業園區等，多個不相關的公司共用同一地址是常見的。這時若公司名稱不同，大概率是 False Positive。
   - 如果是「ShipTo (出貨地址)」比對，即使公司名字不同，若出貨的物理地址完全一致（特別是貨代倉庫或敏感地址），則具備高度的轉運合規風險，請判定為 High。
3. **關聯性與別名解析 (Alias & Associations)**：
   - 檢查黑名單 Detail (Content) 中是否有列出 alias (別名) 或轉投資股權結構，看看查詢原名是否正是其別名之一。

請做出最終的合規審計判定：
- **High (同一實體 / 高風險轉運)**: 
  - **兩者均相同**: 核心名稱完全對上，且物理地址高度相符。
  - **地址相同（名稱不同）**: 核心名稱不同，但出貨地址/物理地址完全一致（存在高度白手套轉運/繞道合規風險）。
- **Medium (關聯企業)**: 
  - **名稱相同（地址不同）**: 核心名稱相同，但位於不同國家/城市或分支機構。
  - **別名/轉投資命中**: 查詢名稱命中黑名單 Detail 中記載的別名(Alias)或轉投資關聯企業。
- **Low (低風險/懷疑)**: 有些微相似度，但無法判定。
- **False Positive (確定誤判)**: 字面相似或因共享園區大樓被篩出，但兩者名稱品牌毫不相干，且非同一實體。

針對風險判定，僅有 **High** 與 **Medium** 需要歸類「風險判定依據」(risk_factor)，可選標準值如下：
- 若判定為 **High**，risk_factor 只能是: "兩者均相同" 或 "地址相同（名稱不同）"
- 若判定為 **Medium**，risk_factor 只能是: "名稱相同（地址不同）" 或 "別名/轉投資命中"
- 若判定為 **Low** 或 **False Positive**，risk_factor 請直接填寫 "" (空字串，不需要判定依據)

請「僅」回覆以下標準的 JSON 格式（不要包含任何前後引導廢話或 markdown 的 ` ```json ` 標記，純 JSON 內容，確保可以被 json.loads 解析）：
{{
  "reasoning": "繁體中文的優雅流暢分析推理過程，列出名稱、地址及關聯性的具體論據（注意：切勿在 reasoning 的字串內部使用未經轉義的雙引號，若需用引號請使用單引號或是 \\\"）",
  "match_level": "High" / "Medium" / "Low" / "False Positive",
  "risk_factor": "兩者均相同" / "地址相同（名稱不同）" / "名稱相同（地址不同）" / "別名/轉投資命中" / ""
}}
"""
    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=1000
        )
        resp_text = response.choices[0].message.content.strip()
        
        # 嘗試解析 JSON 內容
        # 移除可能夾帶的 ```json  ``` 區塊
        clean_json_str = resp_text
        if "```" in resp_text:
            match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", resp_text, re.DOTALL)
            if match:
                clean_json_str = match.group(1)
            else:
                # 簡單清理
                clean_json_str = resp_text.replace("```json", "").replace("```", "").strip()
        
        # 硬性防禦機制：如果模型在 JSON 中使用了非法的未轉義雙引號
        # 這邊我們先把常見的 Markdown 程式碼區塊標記徹底移除
        clean_json_str = clean_json_str.strip()
        
        try:
            judgment = json.loads(clean_json_str)
        except json.JSONDecodeError:
            # 進一步容錯：很多時候是 reasoning 的雙引號或者是換行符號問題。
            # 我們嘗試用更寬鬆的正則表達式把 match_level 與 risk_factor 撈出來
            match_level_found = "Uncertain"
            risk_factor_found = ""
            reasoning_found = "解析失敗，改由正則提取"
            
            lvl_match = re.search(r'"match_level"\s*:\s*"([^"]+)"', clean_json_str, re.I)
            rf_match = re.search(r'"risk_factor"\s*:\s*"([^"]+)"', clean_json_str, re.I)
            reason_match = re.search(r'"reasoning"\s*:\s*"(.+?)"\s*,\s*"(?:match_level|risk_factor)"', clean_json_str, re.DOTALL | re.I)
            
            if lvl_match:
                match_level_found = lvl_match.group(1)
            if rf_match:
                risk_factor_found = rf_match.group(1)
            if reason_match:
                reasoning_found = reason_match.group(1).replace('\\"', '"').replace('\n', ' ')
            else:
                # 如果無法定位 reasoning 的結尾，就用原本的內容
                reasoning_found = f"原始生成：{resp_text}"
                
            judgment = {
                "reasoning": reasoning_found,
                "match_level": match_level_found,
                "risk_factor": risk_factor_found
            }
            
        return judgment
    except Exception as e:
        print(f"呼叫 LLM 發生錯誤: {e}. 原始回覆: {resp_text if 'resp_text' in locals() else ''}")
        return {
            "reasoning": f"解析 LLM 失敗：{str(e)}",
            "match_level": "Uncertain",
            "risk_factor": ""
        }

def process_single_pair(pair, idx, total_count):
    """
    包裝單筆比對任務，用於多執行緒併發處理。
    """
    judgment = get_llm_judgment(pair)
    match_level = judgment.get("match_level", "Uncertain")
    raw_risk_factor = judgment.get("risk_factor", "")
    
    # 只有 High 與 Medium 才顯示風險判定依據，Low / False Positive / 其他一律為空字串
    if match_level in ["High", "Medium"]:
        risk_factor = raw_risk_factor if raw_risk_factor else ""
    else:
        risk_factor = ""

    item = {
        "來源檔案": pair["source_file"],
        "條件ID": pair["condition_id"],
        "查詢名稱": pair["query_name"],
        "查詢國家": pair["query_country"],
        "查詢城市": pair["query_city"],
        "查詢地址": pair["query_address"],
        "黑名單ID": pair["party_id"],
        "黑名單名稱": pair["party_name"],
        "黑名單地址": pair["party_address"],
        "黑名單完整資訊": pair["party_content"][:500] + "..." if len(pair["party_content"]) > 500 else pair["party_content"],
        "原XML命中率": f"{pair['xml_percentage']}%",
        "LLM研判等級": match_level,
        "風險判定依據": risk_factor,
        "LLM分析推理理由": judgment.get("reasoning", "無描述")
    }
    return item

def main():
    xml_files = glob.glob("testfile/*_raw.xml")
    if not xml_files:
        print("在 testfile/ 目錄下找不到任何以 _raw.xml 結尾的檔案！")
        return
        
    print(f"找到 {len(xml_files)} 個 raw XML 檔案。開始進行高風險（>=75%）名單提取...")
    
    all_pairs = []
    for xml_path in xml_files:
        pairs = parse_xml_file(xml_path)
        file_name = os.path.basename(xml_path)
        for p in pairs:
            p["source_file"] = file_name
        all_pairs.extend(pairs)
        
    total_count = len(all_pairs)
    print(f"\n提取完成！在所有 XML 檔案中，共找到 {total_count} 筆命中率 >= 75.0% 的高風險候選資料。")
    
    if total_count == 0:
        print("非常安全！所有查詢結果中均無大於或等於 75% 的黑名單結果，無需進入 LLM 審判。")
        return
        
    print(f"現在啟動本地 LLM gemma-4:31B 進行平行（可配置 concurrent={MAX_WORKERS}）深度合規審查...")
    
    audit_results = []
    
    # 2. 多執行緒併發處理 (使用 ThreadPoolExecutor 並配置 tqdm 進度條)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(process_single_pair, pair, idx, total_count): idx for idx, pair in enumerate(all_pairs, 1)}
        
        for future in tqdm(as_completed(futures), total=total_count, desc="LLM Auditing Progress", unit="pairs"):
            try:
                item = future.result()
                audit_results.append(item)
            except Exception as e:
                print(f"處理 Future 發生嚴重錯誤: {e}")
        
    # 3. 輸出成 Excel 報表 (每次產生包含時間戳記的全新檔案，完美避開 PermissionError)
    df = pd.DataFrame(audit_results)
    
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_excel = f"ICP_Audit_Report_{timestamp}.xlsx"
    
    try:
        df.to_excel(output_excel, sheet_name="全審計結果清單", index=False)
        print(f"\n[OK] 原始審計結果已寫入檔案: {os.path.abspath(output_excel)}")
        print("[INFO] 正在自動美化 Excel 並建立「高與中風險審計閱讀版」及「高與中風險清單」分頁...")
        generate_formatted_excel_report(output_excel)
        print(f"[SUCCESS] 全部報表整合完成！最新成果已成功輸出至全新檔案: {os.path.abspath(output_excel)}")
    except Exception as e:
        print(f"[ERROR] 儲存 Excel 時發生非預期錯誤: {e}")
    
    # 4. 列出統計摘要
    print("\n========= 審計結果統計摘要 =========")
    summary_df = df["LLM研判等級"].value_counts().to_frame()
    print(summary_df)
    print("====================================")

if __name__ == "__main__":
    main()
